import json
import csv
import openpyxl
from datetime import datetime
from tabulate import tabulate


class TransactieData():
  data = {'ongesorteerd' : {'transacties': [], 'zoektermen': []}}
  beschermdeCategorieen = ['ongesorteerd', 'alles', '']

  def beginOpnieuw(self):
    self.data = {'ongesorteerd' : {'transacties': [], 'zoektermen': []}}

  def categorieOverzicht(self):
    categorieLijst = sorted(self.data.keys())
    categorieLijst.pop(categorieLijst.index('ongesorteerd'))
    categorieLijst.insert(0, 'ongesorteerd')
    return categorieLijst

  def beschermdeCategorieLijst(self):
    return self.beschermdeCategorieen

  def voegCategorieToe(self, categorie):
    if not type(categorie) == str:
      return
    if categorie.lower() in self.beschermdeCategorieen:
      return
    if categorie.lower() in self.data:
      return
    self.data[categorie.lower()] = {'transacties': [], 'zoektermen': []}

  def verwijderCategorie(self, categorie):
    if categorie in self.beschermdeCategorieen:
      return
    if not categorie in self.data:
      return
    self.data['ongesorteerd']['transacties'] += self.data[categorie]['transacties']
    self.data.pop(categorie)

  def voegZoektermToe(self, zoekterm, categorie):
    if zoekterm == '':
      return
    if categorie.lower() in self.beschermdeCategorieen:
      return
    if categorie.lower() not in self.data:
      return
    self.data[categorie.lower()]['zoektermen'].append(zoekterm.lower())

  def verwijderZoekterm(self, zoekterm, categorie):
    if categorie.lower() in self.beschermdeCategorieen:
      return
    if categorie.lower() not in self.data:
      return
    if zoekterm.lower() not in self.data[categorie.lower()]['zoektermen']:
      return
    self.data[categorie.lower()]['zoektermen'].remove(zoekterm.lower())

  def zoektermOverzicht(self):
    out = ''
    for categorie in sorted(self.data):
      if categorie in self.beschermdeCategorieen:
        continue
      out += f'{categorie}: ['
      for zoekterm in self.data[categorie]['zoektermen']:
        out += f'{zoekterm}, '
      if out[-2] == ',':
        out = out[:-2]
      out += ']\n'
    return out

  def zoektermenOpslaan(self, path):
    saveData = {}
    for categorie in self.data:
      saveData[categorie] = self.data[categorie]['zoektermen']

    with open(path, 'w') as file:
      json.dump(saveData, file, indent=2, sort_keys=True)

  def zoektermenLaden(self, path):
    with open(path) as file:
      loadData = json.load(file)
    
    for categorie in loadData:
      if not type(categorie) == str:
        return
      if categorie.lower() in self.beschermdeCategorieen:
        continue
      if categorie.lower() in self.data:
        self.data[categorie.lower()]['zoektermen'] += loadData[categorie]
      else:
        self.data[categorie.lower()] = {'transacties': [], 'zoektermen': loadData[categorie]}

  def voegTransactieToe(self, datum, bedrag, tegenrekening, omschrijving, commentaar=''):
    self.data['ongesorteerd']['transacties'].append([datum, bedrag, tegenrekening, omschrijving, commentaar])

  def wijzigCommentaar(self, transactieCategorie, transactieIndex, nieuwCommentaar):
    if transactieCategorie in self.data:
      if len(self.data[transactieCategorie]['transacties'])>transactieIndex:
        self.data[transactieCategorie]['transacties'][transactieIndex][4] = nieuwCommentaar

  def wijzigCategorie(self, huidigeCategorie, huidigeIndex, nieuweCategorie):
    if huidigeCategorie in self.data:
      if len(self.data[huidigeCategorie]['transacties'])>huidigeIndex:
        if nieuweCategorie in self.data:
          self.data[nieuweCategorie]['transacties'].append(self.data[huidigeCategorie]['transacties'].pop(huidigeIndex))

  def verwijderAlleTransacties(self):
    for categorie in self.data:
      self.data[categorie]['transacties'] = []

  def sorteerTransacties(self, categorie='alles', sorteerHeader='datum', rev=False):
    headers=['datum', 'bedrag', 'categorie', 'tegenrekening', 'omschrijving', 'commentaar']
    if sorteerHeader in headers:
      if categorie=='alles':
        for categorie in self.data:
          self.sorteerTransacties(categorie, sorteerHeader, rev)
      else:
        if categorie in self.data:
          self.data[categorie]['transacties'] = sorted(self.data[categorie]['transacties'], key=lambda x:x[headers.index(sorteerHeader)], reverse=rev)

  def categoriseerTransactiesMetZoektermen(self):
    for categorie in self.data:
      zoektermen = self.data[categorie]['zoektermen']
      N = len(self.data['ongesorteerd']['transacties'])
      for i in reversed(range(N)):
        if any([x for x in zoektermen if x.lower() in self.data['ongesorteerd']['transacties'][i][2].lower()]) or any([x for x in zoektermen if x.lower() in self.data['ongesorteerd']['transacties'][i][3].lower()]):
          self.data[categorie]['transacties'].append(self.data['ongesorteerd']['transacties'].pop(i))

  def transactieTabel(self, categorie='alles', headers=['datum', 'bedrag', 'categorie', 'tegenrekening', 'omschrijving', 'commentaar'], startDatum=None, eindDatum=None, sorteerHeader='datum', rev=False):
    tabel = []
    if categorie=='alles' :
      for categorie in self.data:
        tabel.extend(self.transactieTabel(categorie, headers, startDatum, eindDatum, sorteerHeader, rev))    
    else:
      if categorie in self.data:
        for x in self.data[categorie]['transacties']:
          datum = x[0]
          if (startDatum and datum<startDatum) or (eindDatum and datum>eindDatum):
            continue
          rij = []
          for hdr in headers:
            if hdr == 'datum':
              rij.append(datetime.strftime(x[0], '%d/%m/%Y'))
            elif hdr == 'bedrag':
              rij.append(x[1])
            elif hdr == 'categorie':
              rij.append(categorie)
            elif hdr == 'tegenrekening':
              rij.append(x[2])
            elif hdr == 'omschrijving':
              rij.append(x[3])
            elif hdr == 'commentaar':
              rij.append(x[4])
          tabel.append(rij)

    if sorteerHeader in headers:
      if sorteerHeader == 'datum':
        tabel = sorted(tabel, key=lambda x:datetime.strptime(x[headers.index('datum')], '%d/%m/%Y'), reverse=rev)
      else:
        tabel = sorted(tabel, key=lambda x:x[headers.index(sorteerHeader)], reverse=rev)
    return tabel

  def transactieOverzicht(self, categorie='alles', headers=['datum', 'bedrag', 'categorie', 'tegenrekening', 'omschrijving', 'commentaar'], startDatum=None, eindDatum=None, sorteerHeader='datum', rev=False):
    tabel = self.transactieTabel(categorie, headers, startDatum, eindDatum, sorteerHeader, rev)
    return tabulate(tabel, headers=headers)

  def transactieLijst(self, categorie='alles', headers=['datum', 'bedrag', 'categorie', 'tegenrekening', 'omschrijving', 'commentaar'], startDatum=None, eindDatum=None, sorteerHeader='datum', rev=False):
    transactieOverzicht = self.transactieOverzicht(categorie, headers, startDatum, eindDatum, sorteerHeader, rev)
    return transactieOverzicht.split('\n')

  def berekenTotalen(self, categorie='alles', startDatum=None, eindDatum=None):
    totaal = 0.
    inTotaal = 0.
    uitTotaal = 0.
    aantal = 0
    if categorie=='alles':
      for categorie in self.data:
        subTotaal, inSubTotaal, uitSubTotaal, subAantal = self.berekenTotalen(categorie, startDatum, eindDatum)
        totaal += subTotaal
        inTotaal += inSubTotaal
        uitTotaal += uitSubTotaal
        aantal += subAantal
    else:
      if categorie in self.data:
        for x in self.data[categorie]['transacties']:
          datum = x[0]
          if (startDatum and datum < startDatum) or (eindDatum and datum > eindDatum):
            continue 

          bedrag=x[1]
          totaal += bedrag
          aantal+=1
          if bedrag > 0:
            inTotaal += bedrag
          else:
            uitTotaal += bedrag
    return totaal, inTotaal, uitTotaal, aantal

  def transactiesExporterenCsv(self, path):
    headers=['datum', 'bedrag', 'categorie', 'tegenrekening', 'omschrijving', 'commentaar']
    tabel = self.transactieTabel('alles')
    with open(path, 'w', newline='') as f:
      writer = csv.writer(f) 
      writer.writerow(headers)
      writer.writerows(tabel)
    return

  def transactiesImporterenCsv(self, path):
    with open (path, 'r', newline='') as f:
      reader = csv.reader(f)
      next(reader)
      for row in reader:
        datum = datetime.strptime(row[0], '%d/%m/%Y')
        bedrag = float(row[1])
        categorie = row[2]
        tegenrekening = row[3]
        omschrijving = row[4]
        commentaar = '' if len(row)<6 else row[5]
        if categorie.lower() not in self.data:
          self.data[categorie.lower()] = {'transacties': [], 'zoektermen': []}
        self.data[categorie.lower()]['transacties'].append([datum, bedrag, tegenrekening, omschrijving, commentaar])

  def transactiesImporterenTriodos(self, path):
    wb = openpyxl.load_workbook(path)
    ws = wb._sheets[0]
    for i in range(2, ws.max_row):
      datum = ws.cell(row=i, column=1).value
      bedrag = ws.cell(row=i, column=3).value
      debCred = ws.cell(row=i, column=4).value
      tegenrekening = ws.cell(row=i, column=5).value
      omschrijving = ws.cell(row=i, column=8).value
      if not type(datum).__name__ == 'datetime':
        continue
      if not omschrijving:
        omschrijving = ''
      if not tegenrekening:
        tegenrekening = ''
      if debCred=='Debet':
        bedrag*=-1

      self.voegTransactieToe(datum, bedrag, tegenrekening, omschrijving)
    return

  def transactiesImporterenAsn(self, path):
    with open (path, 'r', newline='') as f:
      reader = csv.reader(f)
      for row in reader:
        datum = datetime.strptime(row[0], '%d-%m-%Y')
        bedrag = float(row[10])
        tegenrekening = row[3]
        omschrijving = row[17]
        self.voegTransactieToe(datum, bedrag, tegenrekening, omschrijving)
    return

  def datumRange(self):
    firstDate= datetime.now()
    lastDate = datetime.now()
    
    first = True
    for categorie in self.data:
      for transactie in self.data[categorie]['transacties']:
        datum = transactie[0]
        if first:
          firstDate = datum
          lastDate = datum
          first = False
        else:
          if datum > lastDate:
            lastDate = datum
          elif datum < firstDate:
            firstDate = datum
    return firstDate, lastDate
